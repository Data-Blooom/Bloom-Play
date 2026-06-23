import os
import time
from datetime import datetime
from openpyxl import Workbook, load_workbook


class SystemLogger:
    def __init__(self, data_provider, interval=300, file_name=None):
        self.data_provider = data_provider
        self.interval = interval

        base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        self.file_name = file_name or os.path.join(base_dir, "bloomplay_log.xlsx")

        print("Logger file path:", self.file_name)

        self.ensure_file()

    def ensure_file(self):
        if not os.path.exists(self.file_name):
            wb = Workbook()
            ws = wb.active
            ws.title = "Logs"

            ws.append([
                "Time",
                "CPU %",
                "CPU Temp",
                "GPU %",
                "GPU Temp",
                "RAM Used",
                "RAM Total",
                "VRAM Used",
                "VRAM Total",
                "Ping"
            ])

            wb.save(self.file_name)

    def log_snapshot(self, label="RUN"):
        try:
            data = self.data_provider()
            gpu = data.get("gpu") or {}

            row = [
                f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ({label})",
                data.get("cpu", 0),
                data.get("cpu_temp", 0),
                gpu.get("usage", 0),
                gpu.get("temp", 0),
                data["ram"]["used"],
                data["ram"]["total"],
                gpu.get("vram_used", 0),
                gpu.get("vram_total", 0),
                data["network"]["ping"]
            ]

            wb = load_workbook(self.file_name)
            ws = wb.active
            ws.append(row)
            wb.save(self.file_name)

            print("Logged:", label)

        except Exception as e:
            print("Logger error:", e)

    def start(self):
        self.log_snapshot(label="START")

        while True:
            time.sleep(self.interval)
            self.log_snapshot(label="AUTO")